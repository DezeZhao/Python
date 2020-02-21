# -*- coding: utf-8 -*-
# @Time    : 2020/2/3 16:28
# @Software: PyCharm
# @File    : test1.py
# @Author  : DezeZhao
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# 第一种方法读写xlsx文件
'''--------------------------------------------------------------------------'''
"""write Excel.xlsx"""
wb = Workbook()
ws = wb.active  # 激活 worksheet

'''
# ws.title = "abc"  # sheet名称
# print(wb.get_sheet_names())  # 得到所有sheet的名字
# ws1=wb.create_sheet(title="sheet1")
# for i in range(1, 10):
#     for j in range(1, 10):
#         ws.cell(i, j, 'AA')
# ws['A1']=43  # 赋值方式

# ws1.append(['张三', 80])#从数据下面追加一行
# ws1.append(['李四', 90])
'''

# 根据列的数字返回字母
print(get_column_letter(2))  # B
# 根据字母返回列的数字
print(column_index_from_string('D'))  # 4

sheet_data = [
    ['index', 'data1', 'data2'],
    [1, 2, 3],
    [2, 3, 4],
    [3, 4, 5],
    [4, 5, 6],
    [5, 6, 7]
]

for i in range(1, 7):
    for j in range(1, 4):
        # ws.cell(i, j, sheet_data[i-1][j-1])
        ws['%c%d' % ((ord('A') - 1 + j), i)] = sheet_data[i - 1][j - 1]  # 同上

'''
for cell in ws['A1:C6']:
    print(cell)
for row_cell in ws['A1':'C6']:
    for cell in row_cell:
        print(cell)
# <Cell 'Sheet'.B1>
'''

wb.save('b.xlsx')

'''-----------------------------------------------------------------------'''
"""read excel.xlsx"""

lwb = load_workbook('b.xlsx')
lws = lwb.active

# 写入值，最大行列值
'''
# x=lws['B3'].value
# print(x)
# print(ls.max_column,ls.max_row)
'''

# 二维list存放读取的数据
dim2_list = []
print('---二维list存放读取的数据---')
for row in lws.rows:
    dim1_list = []
    for cell in row:
        dim1_list.append(cell.value)
    dim2_list.append(dim1_list)
print(dim2_list)

'''
print('---按行获取单元格的值---')
for row in lws.rows:  # <class 'generator'>
    # lws.rows是生成器类型，不能使用索引，故将其转换为list类型即可使用索引来获取某一行的数据
    for cell in row:
        print(cell.value)  # A1 A2 A3 B1 B2 B3...
'''

# 按列
'''
print('---按列获取单元格的值---')
for column in lws.columns:
    for cell in column:
        print(cell.value)
'''

print('---获取单个单元格的值---')
print(lws['A1'].value)
print(lws[1][0].value)  # 这里也是A1值,行索引从1、列索引从0算起

print('---获取单列的所有值---')
for cell in lws['A']:
    print(cell.value)

print('---获取多列的值(通过切片)---')
# 如果范围比实际大，如实际只有AB两列，指定A:B，则获取结果返回None，并且C列会被后面.columns和.rows获取到
for column in lws['A:B']:
    for cell in column:
        print(cell.value)

# 获取某行数据
'''
list1 = []
print('---获取某行的值---')
for cell in list(lws.rows)[1]:  # 第2行数据
    list1.append(cell.value)
'''
print('---获取某行的值---')
for cell in lws[1]:
    print(cell.value)

print('---获取多行的值(通过切片)---')
for row in lws[1:2]:
    for cell in row:
        print(cell.value)

# 行转换为列
'''
# 使用zip()函数将list进行打包为元组的列表
print(*sheet_data)  # 将list,dict,np.array,tuple进行拆分成单个元素
zipped = zip(*sheet_data)
print(list(zipped))
zip(*zipped)  # 解压恢复
# [('index', 1, 2, 3, 4, 5), ('name1', 2, 3, 4, 5, 6), ('name2', 3, 4, 5, 6, 7)]
'''
